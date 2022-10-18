import tkinter as tk
import time
from tkinter import *
from tkinter import filedialog, messagebox
from datetime import datetime
import openpyxl
import os
from openpyxl import load_workbook
from PIL import ImageTk, Image




class FirstPage(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.configure(bg='floral white')


        self.bg = ImageTk.PhotoImage(Image.open("Tiger.png"))

        self.canvas1 = Canvas(self, width=400, height=400)
        self.canvas1.create_image(200, 200, image=self.bg)
        self.canvas1.place(x=880, y=180)


        localtime = time.asctime(time.localtime(time.time()))

        label = tk.Label(self, text="RESTAURANT\n MANAGEMENT SYSTEM", bg="floral white",
                         font=("Lucida Handwriting", 35))
        label.place(x=380, y=10)

        localtime = tk.Label(self, text=localtime, bg="floral white", font=("Freestyle Script", 20))
        localtime.place(x=590, y=135)





        Button = tk.Button(self, text="TABLE 1", width=12, height=4, bg="PeachPuff2", font=("Lucida Handwriting", 20),
                           command=lambda: [controller.show_frame(Table),
                                            Table.labels(self, 1),
                                            Table.save_pn(self, 1),
                                            Table.show(self, 1)])
        Button.place(x=50, y=200)

        Button = tk.Button(self, text="TABLE 2", width=12, height=4, bg="PeachPuff2", font=("Lucida Handwriting", 20),
                           command=lambda: [controller.show_frame(Table),
                                            Table.labels(self, 2),
                                            Table.save_pn(self, 2),
                                            Table.show(self, 2)])
        Button.place(x=300, y=200)

        Button = tk.Button(self, text="TABLE 3", width=12, height=4, bg="PeachPuff2", font=("Lucida Handwriting", 20),
                           command=lambda: [controller.show_frame(Table),
                                            Table.labels(self, 3),
                                            Table.save_pn(self, 3),
                                            Table.show(self, 3)])
        Button.place(x=550, y=200)



        Button = tk.Button(self, text="TABLE 4", width=12, height=4, bg="PeachPuff2", font=("Lucida Handwriting", 20),
                           command=lambda: [controller.show_frame(Table),
                                            Table.labels(self, 4),
                                            Table.save_pn(self, 4),
                                            Table.show(self, 4)])
        Button.place(x=50, y=420)

        Button = tk.Button(self, text="TABLE 5", width=12, height=4, bg="PeachPuff2", font=("Lucida Handwriting", 20),
                           command=lambda: [controller.show_frame(Table),
                                            Table.labels(self, 5),
                                            Table.save_pn(self, 5),
                                            Table.show(self, 5)])
        Button.place(x=300, y=420)

        Button = tk.Button(self, text="TABLE 6", width=12, height=4, bg="PeachPuff2", font=("Lucida Handwriting", 20),
                           command=lambda: [controller.show_frame(Table),
                                            Table.labels(self, 6),
                                            Table.save_pn(self, 6),
                                            Table.show(self, 6)])
        Button.place(x=550, y=420)




class Table(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.configure(bg='floral white')
        localtime = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')[:-3]

        filepath = "excel.xlsx"
        if not os.path.exists(filepath):
            self.wb = openpyxl.Workbook()
            self.sheet = self.wb.active
            heading = ["Masa No:","Döner", "Kebap"]
            self.sheet.append(heading)
            self.wb.save(filepath)





        Button = tk.Button(self, text="Home",bg="PeachPuff2", font=("Arial", 15),
                           command=lambda: [controller.show_frame(FirstPage),
                                            self.delete(), self.teser()])
        Button.place(x=100, y=650)

        self.food =tk. Label(self,text="Foods ", fg="black",bg="PeachPuff2", font=("Helvetica", 28))
        self.food.place(x=65, y=120)



        self.fish = Label(self,text="Fish  ", fg="Red",bg='floral white',  font=("Helvetica", 22))
        self.fish.place(x=65, y=200)

        self.kebap = Label(self,text="Kebap ", fg="Red",bg='floral white',  font=("Helvetica", 22))
        self.kebap.place(x=65, y=260)

        self.chicken = Label(self,text="Chicken  ", fg="Red",bg='floral white',  font=("Helvetica", 22))
        self.chicken.place(x=65, y=320)

        self.drink = Label(self,text="Drinks ", fg="black",bg="PeachPuff2",  font=("Helvetica", 28))
        self.drink.place(x=65, y=380)

        self.cola = Label(self,text="Cola ", fg="Red",bg='floral white',  font=("Helvetica", 22))
        self.cola.place(x=65, y=460)

        self.water = Label(self,text="Water ", fg="Red",bg='floral white',  font=("Helvetica", 22))
        self.water.place(x=65, y=580)

        self.coffee = Label(self,text="Coffee ",fg="Red",bg='floral white',  font=("Helvetica", 22))
        self.coffee.place(x=65, y=520)





        self.wb = load_workbook(filename="excel.xlsx")
        self.ws = self.wb["Sheet"]



        self.entry1=Label(self,bg='floral white',  width=5,font=("Helvetica", 22))
        self.entry1.place(x=200, y=200)
        self.entry1_spinbox=Spinbox(self, from_ = 0, to = 50, font=("Arial", 22), width=5)
        self.entry1_spinbox.place(x=250,y=200)


        self.entry2=Label(self,bg='floral white', width=5,font=("Helvetica", 22))
        self.entry2.place(x=200,y=260)
        self.entry2_spinbox=Spinbox(self,from_=0,to=50,font=("Arial", 22),width=5)
        self.entry2_spinbox.place(x=250,y=260)

        self.entry3=Label(self,bg='floral white', width=5,font=("Helvetica", 22))
        self.entry3.place(x=200,y=320)
        self.entry3_spinbox=Spinbox(self,from_=0,to=50,font=("Arial", 22),width=5)
        self.entry3_spinbox.place(x=250,y=320)



        self.entry4=Label(self,bg='floral white', width=5,font=("Helvetica", 22))
        self.entry4.place(x=200,y=460)
        self.entry4_spinbox=Spinbox(self,from_=0,to=50,font=("Arial", 22),width=5)
        self.entry4_spinbox.place(x=250,y=460)

        self.entry5=Label(self,bg='floral white', width=5,font=("Helvetica", 22))
        self.entry5.place(x=200,y=520)
        self.entry5_spinbox=Spinbox(self,from_=0,to=50,font=("Arial", 22),width=5)
        self.entry5_spinbox.place(x=250,y=520)

        self.entry6=Label(self,bg='floral white', width=5,font=("Helvetica", 22))
        self.entry6.place(x=200,y=580)
        self.entry6_spinbox=Spinbox(self,from_=0,to=50,font=("Arial", 22),width=5)
        self.entry6_spinbox.place(x=250,y=580)





        self.button = tk.Button(self,text="Total", width=10, height=2 ,bg="PeachPuff2", fg="black", font=("Helvetica", 18), command=self.getplus)
        self.button.place(x=750,y=220)

        self.button = tk.Button(self,text="Receipt", width=8, height=2 , fg="black",bg="PeachPuff2", font=("Helvetica", 18), command=self.getReceipt)
        self.button.place(x=750,y=320)

        self.button =tk.Button(self,text="Save\n Receipt", width=8, height=2 , fg="black",bg="PeachPuff2", font=("Helvetica", 18), command=self.save)
        self.button.place(x=900,y=320)

        self.button =tk.Button(self,text="Reset", width=8, height=2 , fg="black",bg="PeachPuff2", font=("Helvetica", 18), command=self.reset)
        self.button.place(x=1050,y=320)

        self.button =tk.Button(self,text="Save \n Excel", width=8, height=2 , fg="black",bg="PeachPuff2", font=("Helvetica", 18),
                               command=self.save_excel)
        self.button.place(x=1200,y=320)




        self.textreceipt=Text(self,font=('arial', 12, 'bold'), bd=3, width=60, height=14)
        self.textreceipt.place(x=750, y=420)
        self.filepath = "excel.xlsx"

    def labels(self,i):
        global label
        t_no="Table No:"+str(i)
        label=Label(text=t_no,bg='floral white',  font=("Helvetica", 28))
        label.place(x=200,y=20)
    def show(self,sn):
        global fish
        global kebap
        global chicken
        global cola
        global coffee
        global water

        self.wb = load_workbook(filename="excel.xlsx")
        self.ws = self.wb["Sheet"]

        a = "B" + str(sn+1)
        b = "C" + str(sn+1)
        c = "D" + str(sn + 1)
        d = "E" + str(sn + 1)
        e = "F" + str(sn + 1)
        f = "G" + str(sn + 1)

        fish = self.ws[a].value
        kebap = self.ws[b].value
        chicken = self.ws[c].value
        cola = self.ws[d].value
        coffee = self.ws[e].value
        water = self.ws[f].value


        fish = Label(text=fish, fg="Red", bg='floral white', font=("Helvetica", 22))
        fish.place(x=380, y=200)

        kebap = Label(text=kebap, fg="Red",bg='floral white',  font=("Helvetica", 22))
        kebap.place(x=380, y=260)

        chicken = Label(text=chicken, fg="Red",bg='floral white',  font=("Helvetica", 22))
        chicken.place(x=380, y=320)

        cola = Label(text=cola, fg="Red",bg='floral white',  font=("Helvetica", 22))
        cola.place(x=380, y=460)

        coffee = Label(text=coffee, fg="Red",bg='floral white',  font=("Helvetica", 22))
        coffee.place(x=380, y=520)

        water = Label(text=water, fg="Red",bg='floral white',  font=("Helvetica", 22))
        water.place(x=380, y=580)

    def delete(self):
        label.destroy()
        fish.destroy()
        kebap.destroy()
        chicken.destroy()
        cola.destroy()
        coffee.destroy()
        water.destroy()


    def save(self):
        url = filedialog.asksaveasfile(mode='w+', defaultextension='.txt')
        bill_data = self.textreceipt.get(1.0, END)
        url.write(bill_data)
        url.close()
        messagebox.showinfo('Info', 'Save is successful!')

    def save_pn(self,i):

        global sn
        sn=i

        print(sn)
        return sn

    def save_excel(self):

        print(sn)

        self.wb = load_workbook(filename="excel.xlsx")
        self.ws = self.wb["Sheet"]
        a = "B" + str(sn + 1)
        b = "C" + str(sn + 1)
        c = "D" + str(sn + 1)
        d = "E" + str(sn + 1)
        e = "F" + str(sn + 1)
        f = "G" + str(sn + 1)

        self.ws[a].value = self.entry1_spinbox.get()
        self.ws[b].value = self.entry2_spinbox.get()
        self.ws[c].value = self.entry3_spinbox.get()
        self.ws[d].value = self.entry4_spinbox.get()
        self.ws[e].value = self.entry5_spinbox.get()
        self.ws[f].value = self.entry6_spinbox.get()

        self.wb.save("excel.xlsx")
        self.wb.close()

    def getReceipt(self,*args):
        global billnumber, date

        self.textreceipt.delete(1.0, END)
        date = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')[:-3]

        self.x="MASA 2"
        billnumber = str(self.x)
        self.textreceipt.insert(END,"\t\t Tiger Scholar Restaurant\n\n")
        self.textreceipt.insert(END, billnumber + '\t\t\t' + date +'\n\n')
        self.textreceipt.insert(END, '****************************************************************\n\n')
        self.textreceipt.insert(END, 'Sipariş: :\t\t Adet\t\tFiyat:\n\n')
        self.textreceipt.insert(END, '****************************************************************\n')

        if self.x1 !='0':
            self.textreceipt.insert(END, f'Fish:\t\t{int(self.x1)}\t\t{int(self.x1)*60} TL\n\n')

        if self.x2 !='0':
            self.textreceipt.insert(END, f'Kebap:\t\t{int(self.x2)}\t\t{(int(self.x2)*45)} TL\n\n')
        if self.x3 !='0':
            self.textreceipt.insert(END, f'Chicken:\t\t{int(self.x3)}\t\t{(int(self.x3)*50)} TL\n\n')

        if self.x4 !='0':
            self.textreceipt.insert(END, f'Chicken:\t\t{int(self.x4)}\t\t{(int(self.x4)*10)} TL\n\n')
        if self.x5 !='0':
            self.textreceipt.insert(END, f'Chicken:\t\t{int(self.x5)}\t\t{(int(self.x5)*20)} TL\n\n')
        if self.x6 !='0':
            self.textreceipt.insert(END, f'Chicken:\t\t{int(self.x6)}\t\t{(int(self.x6)*5)} TL\n\n')




        self.textreceipt.insert(END, '****************************************************************\n')
        self.textreceipt.insert(END, f'KDV :\t\t\t\t{(self.toplam*8)/100} TL\n\n')
        self.textreceipt.insert(END, f'TOPLAM TUTAR:\t\t\t\t{self.toplam} TL\n\n')
        self.textreceipt.insert(END,"Kasiyer:\t\t\t\t Mehmet Alim")

    def reset(self):
        self.textreceipt.delete(1.0, END)
        self.entry1_spinbox.delete(0,END)
        self.entry2_spinbox.delete(0,END)
        self.entry3_spinbox.delete(0, END)
        self.entry4_spinbox.delete(0, END)
        self.entry5_spinbox.delete(0, END)
        self.entry6_spinbox.delete(0, END)


        self.label2.destroy()


    def teser(self):
        self.textreceipt.delete(1.0, END)
        self.entry1_spinbox.delete(0,END)
        self.entry2_spinbox.delete(0,END)
        self.entry3_spinbox.delete(0, END)
        self.entry4_spinbox.delete(0, END)
        self.entry5_spinbox.delete(0, END)
        self.entry6_spinbox.delete(0, END)

    def getplus(self, *args):
        self.x1 = self.entry1_spinbox.get()
        self.x2 = self.entry2_spinbox.get()
        self.x3 = self.entry3_spinbox.get()
        self.x4 = self.entry4_spinbox.get()
        self.x5 = self.entry5_spinbox.get()
        self.x6 = self.entry6_spinbox.get()



        self.toplam = int(self.x1) * 60 + int(self.x2) * 45+int(self.x3) * 50+int(self.x4) * 10+int(self.x5) * 20+int(self.x6) * 5
        self.label2 = Label(self,bg='floral white',  text=(str(self.toplam) + ' TL'), font=('helvetica', 25, 'bold'))
        self.label2.place(x=950, y=235)

class Application(tk.Tk):
    def __init__(self, *args, **kwargs):
        tk.Tk.__init__(self, *args, **kwargs)

        window = tk.Frame(self)
        window.pack()
        self.configure(bg='white')

        window.grid_rowconfigure(0, minsize=750)
        window.grid_columnconfigure(0, minsize=1400)

        self.frames = {}
        for F in (FirstPage, Table):
            frame = F(window, self)
            self.frames[F] = frame
            frame.grid(row=0, column=0, sticky="nsew")

        self.show_frame(FirstPage)

    def show_frame(self, page):
        frame = self.frames[page]
        frame.tkraise()
        self.title("Uygulama")

app = Application()
app.geometry("1400x790+10+10")
app.mainloop()
